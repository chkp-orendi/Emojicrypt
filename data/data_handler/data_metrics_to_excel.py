import os
import json
import openpyxl
import statistics

data_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'testers','metrics','2024-07-23_19_54_22.206368-metrics.json'))



def write_to_excel(data):
    # Create a workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active


    row = 2
    for obfuscator in data:
        ws[f'A{row}'] = obfuscator[0]
        question_sim_avrage = 0
        question_sim_mean = []
        answer_sim_avrage = 0
        answer_sim_mean = []
        precentange_of_words_changed = 0
        for details in obfuscator[1]:
            if (obfuscator[0] == "FakeObfuscator") and (details[3][1]<0.999):
                print(details[0])
            ws[f'B{row}'] = details[3][0]  # Question similarity
            ws[f'C{row}'] = details[4]     # Answer similarity
            ws[f'D{row}'] = details[3][1]  # % of words from list
            ws[f'E{row}'] = details[0]     # Obfuscated Question
            ws[f'F{row}'] = details[1]     # Obfuscated Answer
            ws[f'G{row}'] = details[2]     # Deobfuscated Answer
            question_sim_avrage += details[3][0]
            question_sim_mean.append(details[3][0])
            answer_sim_avrage += details[4]
            answer_sim_mean.append(details[4])
            precentange_of_words_changed += details[3][1]
            row += 1
        ws[f'A{row}'] = "Question similarity Avrage"
        ws[f'B{row}'] = "Question similarity Mean"
        ws[f'C{row}'] = "Answer similarity Avrage"
        ws[f'D{row}'] = "Answer similarity Mean"
        ws[f'E{row}'] = "Precentange of words changed"
        row += 1
        ws[f'A{row}'] = question_sim_avrage/len(question_sim_mean)
        ws[f'B{row}'] = statistics.mean(question_sim_mean)  # Corrected line
        ws[f'C{row}'] = answer_sim_avrage/len(answer_sim_mean)
        ws[f'D{row}'] = statistics.median(answer_sim_mean)
        ws[f'E{row}'] = precentange_of_words_changed/len(obfuscator[1])
        row += 1
    # Save the workbook
    wb.save("test_24_07.xlsx")

def main():
    print("starting")

    with open(data_path) as file:
        data = json.load(file)
        print("opened")

    write_to_excel(data)
        
    
    
if __name__ == "__main__":
    main()
    