import os
import sys
import json
import openpyxl

data_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'testers','metrics','2024-07-23_19_54_22.206368-metrics.json'))



def write_to_excel(data):
    # Create a workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active


    row = 2
    for index,obfuscator in data[0]:
        ws[f'A{row}'] = obfuscator
        for details in data[0][index]:
            ws[f'C{row}'] = details[3][0]  # Question similarity
            ws[f'B{row}'] = details[4][0]  # Answer similarity
            ws[f'D{row}'] = details[3][1]  # % of words from list
            ws[f'E{row}'] = details[0][0]  # Obfuscated Question
            ws[f'F{row}'] = details[1][0]  # Obfuscated Answer
            ws[f'G{row}'] = details[2][0]  # Deobfuscated Answer
        row += 1

    # Save the workbook
    wb.save("output.xlsx")

def main():
    print("starting")

    with open(data_path) as file:
        data = json.load(file)
        print("opened")

    for obfuscator in data[0]:
        print(data[0]["WrongObfuscator"][0][0])
    


if __name__ == "__main__":
    main()
    